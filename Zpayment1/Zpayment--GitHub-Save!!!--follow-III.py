# -*- coding:utf-8 -*-     （  ）
# ****************************************************************
# 程序性质：Excel xls/xlsx 处理
# 功能：从付款信息大表读入，按要求分别存到付款文件中
#
# 使用方法： 1. 在工作文件夹建立Templates子文件夹，先填好里面模板
#           2. 运行此.py程序
# 输出： 全部输出到与Templates并列的output文件夹里
# 注意事项： 1. 注意：设置 各文件夹路径、模板文件名、列名与本程序匹配。不改变的话可直接运行。'
#           2. 也要注意：Excel标题行在第几行，等等。
# *****************************************************************
# auth__ = 'Larry Zhang' # 2020-11-24 1808
# v1.00     2020-11-24      不同的乙方，这版全，但日期都留空了（因合同没签完）
# v1.01     2020-12-5       乙方相同的才放在一起。乙方名称、银行放到信息大表里了。
# v1.02     2020-12-5 900   使用当前路径下的Templates文件夹，找不到的话再用绝对路径。使用pathlib

import sys
import os
from pathlib import Path, PureWindowsPath
import pandas as pd
# import re
# import datetime, time
# import numpy as np    # 20191102 np.NaN用这个
# import xlrd.xldate
# from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
# from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter  # , column_index_from_string  # 不是from .cells里，改为.utils了


def isvvalue(o1):   # 有效值；各种空值返回False, 0返回True
    if pd.isna(o1):
        return False
    elif isinstance(o1, str):
        if o1.rstrip().lstrip():
            return True
        else:
            return False
    else:
        return True


def chn_date(np_datetime):
    if isvvalue(np_datetime):   # if np_datetime:遇到NaN会判为成立！       # not pd.isna(np_datetime): # isna把' '判为True
        return pd.to_datetime(np_datetime).strftime( '%Y年%m月%d日'.encode('unicode_escape').decode('utf8')).encode('utf-8').decode('unicode_escape')
    else:
        return ''


def a_contract_prt(outtable_str, i, output_path):
    try:
        wb = load_workbook(templ_path / outtable_str)  # encoding='gb18030'不像csv模块，没有这个encoding。
    except Exception as e:
        print(f"出现错误！！！文件【{templ_path / outtable_str}】: \n【{repr(e)}】！")  # e.message str(e)
        while input("回车退出...") != '':
            pass
        else:
            print("程序已出错并结束。")
        os._exit(-9)  # sys.exit(-9)

    wb.guess_types = True  # 加个这个，写入百分数后就会显示百分数，否则显示小数
    # print( 'wb.get_named_ranges()=', wb.get_named_ranges() )
    sheetNames = wb.sheetnames
    print( 'wb.sheet_names[0]= 【', sheetNames[0], f'】  第【{i}】个。' )
    sh = wb[(sheetNames[0])]   # sh = wb.get_sheet_by_name('Sheet1')
    # sh = wb.active #也可。没有()
    print( sh.title, sh.max_row, sh.max_column )
    print( get_column_letter(sh.max_column) )  # 从1开始

    if '付款申请单' in outtable_str:
        sh["B6"].value = pay_to_co
        sh["B7"].value = dfBig['合同名称'][i].replace('\n', '')
        sh["B8"].value = chn_date(dfBig['合同签订时间'][i])
        sh["G8"].value = dfBig['合同金额\n（元）'][i]
        sh["C11"].value = dfBig['应付金额\n（元）'][i]
        sh["E12"].value = bank_name
        sh["H12"].value = bank_acc
        stmp = "    根据与贵单位签订的《" + dfBig['合同名称'][i].replace('\n', '') + '》，“' + dfBig['付款条款'][i].replace('\n', '') + "”"
        if isvvalue(dfBig['完成成果简介'][i]):
            stmp += "\n    " + dfBig['完成成果简介'][i].rstrip()
        else:
            stmp += "\n    "
        stmp += "按照约定申请支付" + str(dfBig['应付金额\n（元）'][i]) + "元。"
        if isvvalue(dfBig['成果文件'][i]):
            stmp += "\n    出具成果资料清单： \n" + dfBig['成果文件'][i]
        else:
            pass
        sh["B13"].value = stmp

    elif '付款台账' in outtable_str:
        sh["B2"].value = pay_to_co
        sh["B4"].value = dfBig['付款单位'][i]
        sh["C4"].value = dfBig['合同金额\n（元）'][i]
        sh["D4"].value = dfBig['项目'][i]
        sh["E4"].value = dfBig['未付金额\n（元）'][i]
        sh["F4"].value = dfBig['已付金额'][i]
        sh["G4"].value = dfBig['应付金额\n（元）'][i]
        sh["H4"].value = chn_date(dfBig['付款时间'][i])
        sh["I4"].value = dfBig['备注'][i]
    elif '付款审批单' in outtable_str:
        sh["B6"].value = pay_to_co
        sh["B10"].value = dfBig['项目'][i].replace('\n', '')
        sh["B14"].value = dfBig['合同名称'][i].replace('\n', '')
        sh["B13"].value = chn_date(dfBig['合同签订时间'][i])  # str.split(dfBig['合同签订时间'][i].astype(str), 'T')[0]
        sh["G14"].value = dfBig['合同金额\n（元）'][i]
        sh["C17"].value = dfBig['应付金额\n（元）'][i]
        sh["E18"].value = bank_name
        sh["H18"].value = bank_acc
        sh["B19"].value = f"    根据我单位与{pay_to_co}签订的《" + dfBig['合同名称'][i].replace('\n', '') + "》的约定，申请支付" \
                          + str(dfBig['应付金额\n（元）'][i]) + "元。\n    妥否，请领导批示。"

    newfn = os.path.splitext(outtable_str)[0] + dfBig['文件名缩写'][i] + '.xlsx'
    newfn = str(i) + newfn[1:]  # 替换模板文件名第一个字母，比如 ‘x6’ 换为 16
    wb.save(filename=output_path / newfn)
    wb.close()


if __name__ == '__main__':
    main_path0, _ = os.path.split(os.path.abspath(sys.argv[0]))  # 本主程序所在路径
    main_path = Path(main_path0)  # pathlib化

    templ_path = main_path / 'templates'
    if not templ_path.exists():
        tmpP = PureWindowsPath(r'C:\tmp\付款申请——2020-12——IV\templates')  # <----不用本程序所在文件夹下的templates的话，用这个路径!!！！
        templ_path = Path(tmpP)
    else:
        pass
    if not templ_path.exists():
        print(f"出现错误！！！程序所在文件夹下没找到模板文件夹：【{main_path / 'templates'}】，临时模板路径【{templ_path}】也不存在！")
        while input("回车退出...") != '':
            pass
        else:
            print("程序已出错并结束。")
        os._exit(-9)  # sys.exit(-9)

    out_path = templ_path / '../output'
    if not out_path.exists():
            os.mkdir(out_path)
    iptFP = templ_path / 'zzz、付款信息汇总大表.xlsx'
    # 没用了：num_of_row = 5

    # template names for output, they are strings:
    outts1 = 'x1、付款申请单 - .xlsx'
    outts5 = 'x5、付款台账 - .xlsx'
    outts6 = 'x6、付款审批单 - .xlsx'

    outts_hz = 'z7、票据汇总单 - .xlsx'

    if not os.access( iptFP, os.F_OK ):
        print( "访问不了源表(付款信息大表)【%s】 ! please check..." % iptFP )
        os.system('pause')
        os._exit(-1)    # 这个不报错退出。 exit(-1)会报错。sys.exit(-1)

    # ##############读取大表#########################
    print( 'Loading Mapping File【', iptFP, '】，Please wait ...... ' )
    try:
        mapX = pd.ExcelFile( iptFP )
        dfBig = pd.read_excel( mapX, sheet_name=u'大表', header=3 - 1, usecols="A: Q", index_col=u'序号' )  # header默认=0，是第0+1行，=2是指2+1行。
        dfTitle = pd.read_excel( mapX, sheet_name=u'大表', header=None, skiprows=0, nrows=2, dtype=str, usecols="A: I", index_col=None)
    except Exception as e:
        print(f"出现错误！！！文件【{iptFP}】: \n【{repr(e)}】！")
        while input("回车退出...") != '':
            pass
        else:
            print("程序已出错并结束。")
        os._exit(-9)  # sys.exit(-9)

    pay_to_co = dfTitle.iloc[1, 1]
    bank_name = dfTitle.iloc[1, 3]
    bank_acc = dfTitle.iloc[1, 5]
    print( '大表: \n', dfBig, end='\n' )
    print( '标题、单位名称等: \n', dfTitle, end='\n\n' )
    # print( '单位名称', pay_to_co, end='\n\n' )

    # os.chdir( working_path )

    # 票据汇总单
    try:
        wbhz = load_workbook(templ_path / outts_hz)  # encoding='gb18030'不像csv模块，没有这个encoding。
    except Exception as e:
        print(f"出现错误！！！文件【{templ_path / outts_hz}】: \n【{repr(e)}】！")
        while input("回车退出...") != '':
            pass
        else:
            print("程序已出错并结束。")
        os._exit(-9)  # sys.exit(-9)

    wbhz.guess_types = True  # 加个这个，写入百分数后就会显示百分数，否则显示小数
    SNs = dfBig.index.dropna().astype(int)
    for i in SNs:   # range(1, num_of_row + 1):
        # print( 'wbhz.get_named_ranges()=', wbhz.get_named_ranges() )
        sheetNames = wbhz.sheetnames
        print( 'wbhz.get_sheet_names()=', sheetNames )
        sh = wbhz[(sheetNames[0])]   # sh = wbhz.get_sheet_by_name('Sheet1')
        # sh = wbhz.active #也可。没有()
        print( sh.title, sh.max_row, sh.max_column )
        print( get_column_letter(sh.max_column) )  # 从1开始
        sh[f"A{7 + i}"].value = dfBig['项目'][i].replace('\n', '')
        sh[f"F{7 + i}"].value = dfBig['应付金额\n（元）'][i]
    wbhz.save(filename=out_path / outts_hz)
    wbhz.close()

    # 按每个合同输出Excel文件：
    for i in SNs:    # range(1, num_of_row + 1):   # 直接改数吧 range1-13 =12个。 len(dfBig.index)
        a_contract_prt(outts1, i, out_path)
        a_contract_prt(outts5, i, out_path)
        a_contract_prt(outts6, i, out_path)
